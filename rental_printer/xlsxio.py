#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from datetime import datetime


class XlsxIO(object):

    def __init__(self, records):
        self.ws = load_workbook(records).active
        self.wl = self.xlsx_to_list(self.ws)
        self.prepare()

    def vali(self, value):
        if value is not None:
            if str(value).strip() != "":
                return True
        return False

    def xlsx_to_list(self, ws):
        worklist = list()
        for row in ws.iter_rows(max_row=ws.max_row,
                                max_col=ws.max_column):
            row_dict = {}
            for cell in row:
                if cell.row != 1:
                    column_head = ws[cell.column+'1'].value
                    cell_value = cell.value
                    if self.vali(column_head) and self.vali(cell_value):
                        row_dict.update({column_head: cell_value})

            worklist.append(row_dict)

        return worklist

    def get_seven(self, money):
        sevenNum = list()
        if money >= 0 and money < 99999.99:
            sevenNum.append(int((money*100) % 10))  # 0.01
            sevenNum.append(int((money*10) % 10))  # 0.1
            sevenNum.append(int(money % 10))  # 1
            if money > 9.99:
                sevenNum.append(int((money/10) % 10))  # 十
                if money > 99.99:
                    sevenNum.append(int((money/100) % 10))  # 百
                    if money > 999.99:
                        sevenNum.append(int((money/1000) % 10))  # 千
                        if money > 9999.99:
                            sevenNum.append(int((money/10000) % 10))  # 万
                        else:
                            sevenNum.append(" ")
                    else:
                        sevenNum.append(" ")
                        sevenNum.append(" ")
                else:
                    sevenNum.append(" ")
                    sevenNum.append(" ")
                    sevenNum.append(" ")
            else:
                sevenNum.append(" ")
                sevenNum.append(" ")
                sevenNum.append(" ")
                sevenNum.append(" ")

        elif money < 0:

            money_plus = money-money-money
            sevenNum.append(int((money_plus*100) % 10))  # 0.01
            sevenNum.append(int((money_plus*10) % 10))  # 0.1
            sevenNum.append(int(money_plus % 10))  # 1
            if money_plus > 9.99:
                sevenNum.append(int((money_plus/10) % 10))  # 十
                if money_plus > 99.99:
                    sevenNum.append(int((money_plus/100) % 10))  # 百
                    if money_plus > 999.99:
                        sevenNum.append(int((money_plus/1000) % 10))  # 千
                        if money_plus > 9999.99:
                            sevenNum.append(int((money_plus/10000) % 10))  # 万
                        else:
                            sevenNum.append("-")
                    else:
                        sevenNum.append("-")
                        sevenNum.append(" ")
                else:
                    sevenNum.append("-")
                    sevenNum.append(" ")
                    sevenNum.append(" ")

            else:
                sevenNum.append("-")
                sevenNum.append(" ")
                sevenNum.append(" ")
                sevenNum.append(" ")

        sevenNum.reverse()
        return sevenNum

    def prepare(self):

        for item in self.wl:
            waterfee = 0.0
            elecfee = 0.0
            cleanfee = 0.0
            netfee = 0.0
            managefee = 0.0
            otherfee = 0.0
            TVfee = 0.0
            housefee = 0
            # datetime prepare:
            if 'Rdate' in item:
                Rdate = item.pop('Rdate')
                if isinstance(Rdate, str):
                    date = datetime.strptime(Rdate, "%Y-%m-%d %H:%M:%S")
                    item.update(
                        {'Rdate_seven': [date.year, date.month, date.day]})

                elif isinstance(Rdate, datetime):
                    item.update(
                        {'Rdate_seven': [Rdate.year, Rdate.month, Rdate.day]})

            # water
            if 'waterthis' in item and \
                    'waterlast' in item and 'waterprice' in item:
                waterdegree = float(item['waterthis'])-float(item['waterlast'])
                item.update({'waterdegree': waterdegree})
                waterfee = waterdegree*float(item['waterprice'])
                item.update({'waterfee': waterfee})
                waterfee_seven = self.get_seven(waterfee)
                item.update({'waterfee_seven': waterfee_seven})

            # electricity
            if 'electhis' in item and \
                    'eleclast' in item and 'elecprice' in item:
                elecdegree = float(item['electhis'])-float(item['eleclast'])
                item.update({'elecdegree': elecdegree})
                elecfee = elecdegree*float(item['elecprice'])
                item.update({'elecfee': elecfee})
                elecfee_seven = self.get_seven(elecfee)
                item.update({'elecfee_seven': elecfee_seven})

            # # house
            if 'housefee' in item:
                housefee = float(item['housefee'])
                housefee_seven = self.get_seven(housefee)
                item.update({'housefee_seven': housefee_seven})

            # # total4
            if 'cleanfee' in item and \
                    'TVfee' in item and 'netfee' in item:
                cleanfee = float(item['cleanfee'])
                TVfee = float(item['TVfee'])
                netfee = float(item['netfee'])
                totel_four_seven = self.get_seven(cleanfee+TVfee+netfee)
                item.update({'totel_four_seven': totel_four_seven})

            # # total5
            if 'managefee' in item and \
                    'otherfee' in item:
                managefee = float(item['managefee'])
                otherfee = float(item['otherfee'])
                totel_five_seven = self.get_seven(managefee+otherfee)
                item.update({'totel_five_seven': totel_five_seven})

            # # total money
            total_money = waterfee + elecfee + cleanfee + \
                netfee + managefee + otherfee + TVfee + housefee
            total_money_seven = self.get_seven(total_money)
            item.update({'total_money_seven': total_money_seven})
            total_money_big = self.get_cbig(total_money)
            item.update({'total_money_big': total_money_big})

    def get_cbig(self, n):
        units = ['', '万', '亿']
        nums = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
        decimal_label = ['角', '分']
        small_int_label = ['', '拾', '佰', '仟']
        int_part, decimal_part = str(int(n)), str(n - int(n))[2:]  # 分离整数和小数部分

        res = []
        if decimal_part:
            res.append(''.join(
                [nums[int(x)] + y for x, y in list(zip(decimal_part,
                                                       decimal_label)) if x != '0']))

        if int_part != '0':
            res.append('圆')
            while int_part:
                small_int_part, int_part = int_part[-4:], int_part[:-4]
                tmp = ''.join([nums[int(x)] + (y if x != '0' else '')
                               for x, y in list(zip(small_int_part[::-1],
                                                    small_int_label))[::-1]])
                tmp = tmp.rstrip('零').replace('零零零', '零').replace('零零', '零')
                unit = units.pop(0)
                if tmp:
                    tmp += unit
                    res.append(tmp)
        return ''.join(res[::-1])

    def style_sheet(self, worksheet):
        thin = Side(border_style="thin", color="000000")
        none = Side(border_style="none")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # for merged_cell in worksheet.merged_cell_ranges:
            # self.style_range(worksheet, merged_cell, border=border)

        for rows1 in worksheet.iter_rows(max_row=1, max_col=20):
            for cell1 in rows1:
                cell1.border = Border(
                    top=none, left=none, right=none, bottom=thin)

        for rows2 in worksheet.iter_rows(min_row=13, max_row=13, max_col=20):
            for cell2 in rows2:
                cell2.border = Border(
                    top=none, left=none, right=none, bottom=thin)

        for rows3 in worksheet.get_squared_range(1, 2, 20, 11):
            for cell3 in rows3:
                cell3.border = Border(
                    top=thin, left=thin, right=thin, bottom=thin)

        for rows4 in worksheet.get_squared_range(1, 14, 20, 23):
            for cell4 in rows4:
                cell4.border = Border(
                    top=thin, left=thin, right=thin, bottom=thin)

    def style_range(self, ws, cell_range,
                    border=Border(), fill=None,
                    font=None, alignment=None):

        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = ws[cell_range]
        if font:
            first_cell.font = font

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            l1 = row[0]
            r1 = row[-1]
            l1.border = l1.border + left
            r1.border = r1.border + right
            if fill:
                for c in row:
                    c.fill = fill

    def output(self, template):
        for item in self.wl:
            workbook = load_workbook(template)
            worksheet = workbook.active
            for cell in worksheet.get_cell_collection():
                if self.vali(cell.value):
                    if cell.value.find("{") == 0:
                        end_point = cell.value.find("}")
                        find_name = cell.value[1:end_point]+"_seven"
                        index = int(cell.value[end_point+1:])-1
                        for key in item:
                            if key == find_name:
                                cell.value = item[key][index]
                    elif cell.value.find("[") == 0:
                        end_point = cell.value.find("]")
                        find_name = cell.value[1:end_point]
                        for key in item:
                            if key == find_name:
                                cell.value = item[key]

            self.style_sheet(worksheet)
            if 'district'in item:
                workbook.save(filename="file/"+str(
                    item['district'])+'.xlsx')
                print(str(item['district'])+" is printed")
